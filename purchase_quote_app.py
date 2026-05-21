
# ============================================================
#  Purchase Quote System – Streamlit + Google Colab
#  Run instructions at bottom of file
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, date
import os
import io
import warnings
warnings.filterwarnings("ignore")

import gspread
from google.oauth2.service_account import Credentials

# ── Google Sheets Configuration ──────────────────────────────
CREDENTIALS_PATH = "credentials.json"
MASTER_SPREADSHEET_ID = "1SgKGzdUwjEzGNiRv-NnDFvw6ngJYvHsg0R6M-Y50ZTA"
DATABASE_SPREADSHEET_ID = "1I6osw4QSpH82SMFFEqSb4CR4SYErQby-HF_EQuQpofs"

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

@st.cache_resource
def get_gspread_client():
    try:
        import json
        
        # 1. Check local file first (to prevent local Streamlit from raising an exception when secrets are empty)
        if os.path.exists(CREDENTIALS_PATH):
            with open(CREDENTIALS_PATH, 'r') as f:
                creds_info = json.load(f)
        # 2. Fall back to Streamlit Secrets (for cloud hosting like streamlit.app)
        else:
            try:
                if "gcp_service_account" in st.secrets:
                    creds_info = dict(st.secrets["gcp_service_account"])
                else:
                    st.error("❌ Missing credentials! Set st.secrets['gcp_service_account'] or provide credentials.json")
                    return None
            except Exception as e:
                st.error("❌ Missing credentials! Please provide credentials.json locally or configure Streamlit Secrets.")
                return None
        
        # AUTO-FIX: The private key often gets messed up with literal '\n' strings
        if 'private_key' in creds_info:
            pk = creds_info['private_key']
            # Replace literal "\n" strings with real newline characters
            pk = pk.replace("\\n", "\n")
            # Remove any accidental extra quotes or spaces
            pk = pk.strip().strip('"').strip("'")
            creds_info['private_key'] = pk

        creds = Credentials.from_service_account_info(creds_info, scopes=SCOPES)
        return gspread.authorize(creds)
    except Exception as e:
        st.error(f"❌ Failed to connect to Google Sheets: {e}")
        return None

gc = get_gspread_client()

@st.cache_data(ttl=600)
def load_gsheet(spreadsheet_id: str, sheet_name: str) -> pd.DataFrame:
    """Load data from a Google Sheet. Try API first, then Public CSV fallback."""
    # 1. Try Google Sheets API (gspread)
    if gc:
        try:
            sh = gc.open_by_key(spreadsheet_id)
            ws = sh.worksheet(sheet_name)
            data = ws.get_all_records()
            if data:
                df = pd.DataFrame(data)
                df.columns = [c.strip() for c in df.columns]
                return df
        except Exception as e:
            st.info(f"💡 Note: API access failed for '{sheet_name}'. Trying public link...")

    # 2. Fallback to Public CSV Export (Requires "Anyone with the link" access)
    # Mapping sheet names to their GIDs (you can find these in the URL gid=xxxx)
    gid_map = {
        "Item Master": "17422179",
        "Vendor Master": "359146087",
        "Location Master": "297795439",
        "Purchase Quotes": "0"
    }
    gid = gid_map.get(sheet_name, "0")
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}"
    
    try:
        df = pd.read_csv(url)
        df.columns = [c.strip() for c in df.columns]
        return df
    except Exception as e:
        st.warning(f"⚠️ Could not load '{sheet_name}': {e}")
        return pd.DataFrame()

def load_master_data():
    item_df = load_gsheet(MASTER_SPREADSHEET_ID, "Item Master")
    vendor_df = load_gsheet(MASTER_SPREADSHEET_ID, "Vendor Master")
    loc_df = load_gsheet(MASTER_SPREADSHEET_ID, "Location Master")
    
    # Process Location Master into a dict for the app
    loc_details = {}
    if not loc_df.empty:
        for _, r in loc_df.iterrows():
            code = str(r.get("Location Code", "")).strip()
            if code:
                loc_details[code] = {
                    "GSTIN": str(r.get("Buyer GSTIN", "")),
                    "Address": str(r.get("Delivery Address", "")),
                    "Contact": f"{r.get('Warehouse In-Charge name', '')} ({r.get('Warehouse In-Charge contact No', '')})"
                }
    return item_df, vendor_df, loc_details

# Load them globally for caching
item_df, vendor_df, LOCATION_DETAILS = load_master_data()

# ── Excel storage helpers ────────────────────────────────────
# ── Google Sheets storage helpers ────────────────────────────
REQUIRED_COLS = [
    "Quote No", "Quote Date", "Location Code", "Purchase Quote raised by", "Vendor Name", "Vendor No",
    "Cash", "Online Transfer", "CDC", "PDC", "Credit Days", 
    "Door Delivery", "Pickup", "Courier", "Expected Delivery Date", "Courier Details",
    "Vendor Address", "Vendor City", "Vendor GST", "Vendor Email",
    "S.No", "ERP Code", "Vendor Item No", "Product Description",
    "Qty", "Price Before GST", "GST %", "Price Inc. GST", "Total (Incl. GST)", "Remarks",
    "Freight Charge", "Status", "Created Date", "Modified Date",
    "Payment Method", "Mode of Delivery",
]

LOCAL_EXCEL_PATH = "Purchase_Quotes_Data.xlsx"

def load_quotes() -> pd.DataFrame:
    # Try Google Sheets first
    df = load_gsheet(DATABASE_SPREADSHEET_ID, "Purchase Quotes")
    if not df.empty:
        # Basic cleanup
        for c in REQUIRED_COLS:
            if c not in df.columns: df[c] = ""
        return df
    
    # Fallback to local Excel
    if os.path.exists(LOCAL_EXCEL_PATH):
        try:
            return pd.read_excel(LOCAL_EXCEL_PATH)
        except:
            pass
    return pd.DataFrame(columns=REQUIRED_COLS)

def save_quotes(new_df: pd.DataFrame, is_edit: bool = False, quote_no: str = None):
    # Enforce perfect column alignment to REQUIRED_COLS
    for col in REQUIRED_COLS:
        if col not in new_df.columns:
            new_df[col] = ""
    new_df = new_df[REQUIRED_COLS]

    # 1. Try Google Sheets Saving
    if gc:
        try:
            sh = gc.open_by_key(DATABASE_SPREADSHEET_ID)
            ws = sh.worksheet("Purchase Quotes")
            
            # Ensure Google Sheet headers match REQUIRED_COLS exactly
            headers = ws.row_values(1)
            if not headers:
                ws.append_row(REQUIRED_COLS)
            elif len(headers) < len(REQUIRED_COLS):
                ws.update('A1', [REQUIRED_COLS])
                
            if is_edit and quote_no:
                all_data = ws.get_all_records()
                full_df = pd.DataFrame(all_data)
                
                # Align existing data columns
                for col in REQUIRED_COLS:
                    if col not in full_df.columns:
                        full_df[col] = ""
                full_df = full_df[REQUIRED_COLS]
                
                full_df = full_df[full_df["Quote No"] != quote_no]
                full_df = pd.concat([full_df, new_df], ignore_index=True)
                ws.clear()
                ws.update([full_df.columns.values.tolist()] + full_df.values.tolist())
            else:
                ws.append_rows(new_df.values.tolist())
            return True
        except Exception as e:
            st.error(f"⚠️ Google Sheets Save Failed: {e}")

    # 2. Fallback to Local Excel
    try:
        full_df = pd.DataFrame(columns=REQUIRED_COLS)
        if os.path.exists(LOCAL_EXCEL_PATH):
            full_df = pd.read_excel(LOCAL_EXCEL_PATH)
        
        if is_edit and quote_no:
            full_df = full_df[full_df["Quote No"] != quote_no]
        
        full_df = pd.concat([full_df, new_df], ignore_index=True)
        full_df.to_excel(LOCAL_EXCEL_PATH, index=False)
        st.info("💾 Saved locally to 'Purchase_Quotes_Data.xlsx' as backup.")
        return True
    except Exception as e:
        st.error(f"❌ Error saving even to local Excel: {e}")
        return False

def next_quote_number(df: pd.DataFrame) -> str:
    if df.empty or "Quote No" not in df.columns:
        return "Q0001"
    existing = df["Quote No"].dropna().tolist()
    nums = []
    for q in existing:
        try:
            nums.append(int(str(q).replace("Q", "")))
        except Exception:
            pass
    nxt = max(nums) + 1 if nums else 1
    return f"Q{nxt:04d}"

# ── Page config ──────────────────────────────────────────────
st.set_page_config(
    page_title="Purchase Quote System",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ───────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

/* Sidebar */
section[data-testid="stSidebar"] {
    background: linear-gradient(160deg, #0f172a 0%, #1e293b 100%);
    color: #e2e8f0;
}
section[data-testid="stSidebar"] * { color: #e2e8f0 !important; }
section[data-testid="stSidebar"] .stRadio label { font-size: 15px; }

/* Metric cards */
div[data-testid="metric-container"] {
    background: linear-gradient(135deg, #1e293b, #0f172a);
    border: 1px solid #334155;
    border-radius: 12px;
    padding: 16px;
    color: #f1f5f9;
}
div[data-testid="metric-container"] label { color: #94a3b8 !important; }
div[data-testid="metric-container"] div[data-testid="stMetricValue"] {
    color: #38bdf8 !important; font-size: 2rem !important; font-weight: 700 !important;
}

/* Section headers */
.section-header {
    background: linear-gradient(90deg, #0ea5e9, #6366f1);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    font-size: 1.4rem;
    font-weight: 700;
    margin-bottom: 12px;
}

/* ERP form card */
.erp-card {
    background: #1e293b;
    border: 1px solid #334155;
    border-radius: 14px;
    padding: 24px;
    margin-bottom: 18px;
}

/* Buttons & Download Buttons */
.stButton > button, .stDownloadButton > button {
    background: linear-gradient(135deg, #0ea5e9, #6366f1);
    color: white !important;
    border: none;
    border-radius: 8px;
    padding: 0.5rem 1.8rem;
    font-weight: 600;
    font-size: 15px;
    transition: opacity .2s;
}
.stButton > button:hover, .stDownloadButton > button:hover { opacity: 0.88; }

/* Table */
.stDataFrame { border-radius: 10px; overflow: hidden; }

/* Success / Warning */
div[data-baseweb="notification"] { border-radius: 10px !important; }

/* Dark background for main area */
/* Ensure disabled text is clearly visible in light/dark mode */
div[data-testid="stTextInput"] input:disabled, 
div[data-testid="stNumberInput"] input:disabled {
    color: #0f172a !important;
    -webkit-text-fill-color: #0f172a !important;
    background-color: #f1f5f9 !important;
    font-weight: 600 !important;
    border: 1px solid #cbd5e1 !important;
}

/* ── Line-item table: scrollable, no overflow clipping ── */
div[data-testid="stHorizontalBlock"] {
    overflow-x: auto;
}

/* Number inputs — right-align, never truncate large values */
div[data-testid="stNumberInput"] input {
    text-align: right !important;
    white-space: nowrap !important;
    min-width: 100px !important;
}

/* Price & Total fields — wider min-width */
div[data-testid="stTextInput"] input {
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}

/* Vendor selectbox — allow long names to wrap inside the dropdown */
div[data-testid="stSelectbox"] div[data-baseweb="select"] div {
    white-space: normal !important;
    word-wrap: break-word !important;
}
div[data-testid="stSelectbox"] ul {
    max-height: 300px !important;
    overflow-y: auto !important;
}

/* Column header strip for line-items */
.li-header {
    display: flex;
    gap: 4px;
    font-weight: 600;
    font-size: 12px;
    padding: 6px 0 4px 0;
    border-bottom: 2px solid #334155;
    margin-bottom: 4px;
    color: #94a3b8;
    text-transform: uppercase;
    letter-spacing: 0.04em;
}
</style>
""", unsafe_allow_html=True)

# ── Session state defaults ───────────────────────────────────
if "role" not in st.session_state:
    st.session_state.role = "Product manager / product team"
if "line_items" not in st.session_state:
    st.session_state.line_items = []
if "edit_quote_no" not in st.session_state:
    st.session_state.edit_quote_no = None
if "page" not in st.session_state:
    st.session_state.page = "Dashboard"
if "quote_submitted" not in st.session_state:
    st.session_state.quote_submitted = False   # duplicate-submit guard
if "submit_validated" not in st.session_state:
    st.session_state.submit_validated = False  # show errors only after submit attempt

# (LOCATION_DETAILS is now loaded dynamically from Google Sheets above)

# ── Sidebar ──────────────────────────────────────────────────
with st.sidebar:
    # Display Logo
    try:
        st.image("logo.png", use_container_width=True)
    except:
        pass
    
    st.markdown("## 📋 Purchase Quote System")
    st.markdown("---")
    # Both roles now use "Create / Edit Quote" as the primary workspace
    pages = ["Dashboard", "Create / Edit Quote", "View Quotes"]
    
    st.session_state.page = st.radio("Navigation", pages, index=pages.index(st.session_state.page) if st.session_state.page in pages else 0)
    st.markdown("---")
    st.caption("v1.0 | Purchase Quote ERP")

# (item_df and vendor_df are already loaded globally above)
quotes_df = load_quotes()

# ── Helper: safe column access ───────────────────────────────
def col(df, *names, default=""):
    for n in names:
        if n in df.columns:
            return df[n]
    return pd.Series([default] * len(df))

# ── PO Excel Export ────────────────────────────────────────────
def export_po_excel(quote_no: str, df: pd.DataFrame) -> io.BytesIO:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"PO_{quote_no}"
    
    dark_blue = PatternFill(start_color="001f497d", end_color="001f497d", fill_type="solid")
    light_blue = PatternFill(start_color="00b4c6e7", end_color="00b4c6e7", fill_type="solid")
    dark_red = PatternFill(start_color="00c00000", end_color="00c00000", fill_type="solid")
    light_orange = PatternFill(start_color="fce4d6", end_color="fce4d6", fill_type="solid")
    
    w_bold = Font(color="FFFFFF", bold=True)
    b_bold = Font(color="000000", bold=True)
    c_align = Alignment(horizontal="center", vertical="center")
    r_align = Alignment(horizontal="right", vertical="center")
    
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    q_df = df[df["Quote No"] == quote_no]
    if q_df.empty: return None
    row1 = q_df.iloc[0]
    
    vendor_name = str(row1.get("Vendor Name", ""))
    v_addr = str(row1.get("Vendor Address", ""))
    v_city = str(row1.get("Vendor City", ""))
    v_gst = str(row1.get("Vendor GST", ""))
    v_email = str(row1.get("Vendor Email", ""))
    
    if (not v_addr or v_addr == "nan") and vendor_name:
        vendor_df = load_master_sheet("Vendor Master")
        if not vendor_df.empty:
            vrow = vendor_df[vendor_df["Name"] == vendor_name] if "Name" in vendor_df.columns else vendor_df[vendor_df["name"] == vendor_name]
            if not vrow.empty:
                vr = vrow.iloc[0]
                v_addr = str(vr.get("address", vr.get("Address", "")))
                v_city = str(vr.get("city", vr.get("City", "")))
                v_gst = str(vr.get("vatRegistrationNo", vr.get("GST", vr.get("GSTIN", ""))))
                v_email = str(vr.get("email", vr.get("Email", vr.get("E-Mail", ""))))
    
    v_addr = v_addr if v_addr != "nan" else ""
    v_city = v_city if v_city != "nan" else ""
    v_gst = v_gst if v_gst != "nan" else ""
    v_email = v_email if v_email != "nan" else ""
    
    # ── PO Header Details ──
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15

    ws.merge_cells('A1:I1')
    ws['A1'] = "SUPREME COMPUTERS INDIA PVT. LTD."
    ws['A1'].font = Font(name='Calibri', size=16, bold=True)
    ws['A1'].alignment = c_align
    
    ws.merge_cells('A2:I2')
    ws['A2'] = "18/18, Majestic Plaza, Narasingapuram Street, Mount Road, Chennai - 600002"
    ws['A2'].font = Font(name='Calibri', size=11)
    ws['A2'].alignment = c_align
    
    loc_code = str(row1.get("Location Code", ""))
    loc_info = LOCATION_DETAILS.get(loc_code, {})
    gstin_val = loc_info.get("GSTIN", "33AAOCS1408H1ZR")
    ws.merge_cells('A3:I3')
    ws['A3'] = f"GSTIN: {gstin_val}  |  Tel: 044-42082020"
    ws['A3'].font = Font(name='Calibri', size=11)
    ws['A3'].alignment = c_align

    logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
    if os.path.exists(logo_path):
        from openpyxl.drawing.image import Image
        img = Image(logo_path)
        img.width = 120
        img.height = 40
        img.anchor = 'A1'
        ws.add_image(img)

    ws.merge_cells('A5:I5')
    ws['A5'] = "PURCHASE ORDER"
    ws['A5'].fill = dark_blue
    ws['A5'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    ws['A5'].alignment = c_align

    def set_cell(r, c1, c2, val, is_label=False):
        cell = ws[f'{c1}{r}']
        cell.value = val
        cell.border = tb
        if is_label:
            cell.font = w_bold
            cell.fill = dark_blue
        else:
            cell.fill = light_orange
            cell.alignment = Alignment(wrap_text=True)
        if c1 != c2:
            ws.merge_cells(f'{c1}{r}:{c2}{r}')
            # Apply border to all merged cells
            for col in range(openpyxl.utils.column_index_from_string(c1), openpyxl.utils.column_index_from_string(c2) + 1):
                ws.cell(row=r, column=col).border = tb

    # Row 7: PO Number & Date
    set_cell(7, 'A', 'B', "PO Number:", True)
    set_cell(7, 'C', 'E', quote_no)
    set_cell(7, 'F', 'G', "PO Date:", True)
    set_cell(7, 'H', 'I', row1.get("Quote Date", ""))

    # Row 8: Vendor Name
    set_cell(8, 'A', 'B', "Vendor Name:", True)
    set_cell(8, 'C', 'I', vendor_name)

    # Row 9: Vendor Address
    set_cell(9, 'A', 'B', "Vendor Address:", True)
    full_addr = v_addr
    if v_city: full_addr += f", {v_city}"
    set_cell(9, 'C', 'I', full_addr)

    # Row 10: Vendor GSTIN & Email
    set_cell(10, 'A', 'B', "Vendor GSTIN:", True)
    set_cell(10, 'C', 'E', v_gst)
    set_cell(10, 'F', 'G', "Vendor Email:", True)
    set_cell(10, 'H', 'I', v_email)

    # Row 11: Vendor No & Payment Terms
    set_cell(11, 'A', 'B', "Vendor No:", True)
    set_cell(11, 'C', 'E', str(row1.get("Vendor No", "")))
    set_cell(11, 'F', 'G', "Payment Terms:", True)
    set_cell(11, 'H', 'I', str(row1.get("Credit Days", "")))

    # Row 12: Location & Payment Method
    set_cell(12, 'A', 'B', "Location:", True)
    set_cell(12, 'C', 'E', loc_code)
    set_cell(12, 'F', 'G', "Payment Method:", True)
    set_cell(12, 'H', 'I', str(row1.get("Payment Method", "")))

    # Row 13: Mode of Delivery & Delivery Date
    md = str(row1.get("Mode of Delivery", ""))
    c_details = str(row1.get("Courier Details", ""))
    if c_details: md += f" ({c_details})"
    set_cell(13, 'A', 'B', "Mode of Delivery:", True)
    set_cell(13, 'C', 'E', md)
    set_cell(13, 'F', 'G', "Delivery Date:", True)
    set_cell(13, 'H', 'I', str(row1.get("Expected Delivery Date", "")))

    # ── Delivery Address Box ──
    ws.merge_cells('A15:I15'); ws['A15'] = "Delivery Address"; ws['A15'].fill = dark_blue; ws['A15'].font = w_bold
    for col in range(1, 10): ws.cell(row=15, column=col).border = tb
    
    ws.merge_cells('A16:I16'); ws['A16'] = loc_info.get("Address", "")
    for col in range(1, 10): ws.cell(row=16, column=col).border = tb
    
    ws.merge_cells('A17:I17'); ws['A17'] = f"Contact: {loc_info.get('Contact', '')}"
    for col in range(1, 10): ws.cell(row=17, column=col).border = tb
        
    headers = ["S.No", "ERP Code", "Vendor Item No.", "Product Description", "Qty", "Price (Excl.)", "GST %", "Total (Incl. GST)", "Remarks"]
    for i, h in enumerate(headers, 1): c = ws.cell(row=19, column=i, value=h); c.fill = dark_red; c.font = w_bold; c.alignment = c_align; c.border = tb
        
    r = 20; sub = 0.0
    for _, l in q_df.iterrows():
        ws.cell(row=r, column=1, value=l.get("S.No", "")).border = tb; ws.cell(row=r, column=2, value=l.get("ERP Code", "")).border = tb
        ws.cell(row=r, column=3, value=l.get("Vendor Item No", "")).border = tb; ws.cell(row=r, column=4, value=l.get("Product Description", "")).border = tb
        qty = float(l.get("Qty", 0) or 0); pb = float(l.get("Price Before GST", 0) or 0)
        pi = float(l.get("Price Inc. GST", 0) or 0); gstp = float(l.get("GST %", 0) or 0)
        line_total = qty * pi
        sub += (qty * pb)
        ws.cell(row=r, column=5, value=qty).border = tb; ws.cell(row=r, column=6, value=pb).border = tb
        ws.cell(row=r, column=7, value=gstp).border = tb; ws.cell(row=r, column=8, value=line_total).border = tb; ws.cell(row=r, column=9, value=l.get("Remarks", "")).border = tb
        for c in range(1, 10): ws.cell(row=r, column=c).fill = light_orange
        r += 1
        
    while r <= 30:
        for c in range(1, 10): ws.cell(row=r, column=c).border = tb
        r += 1
        
    fr = float(row1.get("Freight Charge", 0) or 0)
    ti = sum([float(l.get("Price Inc. GST", 0) or 0) * float(l.get("Qty", 0) or 0) for _, l in q_df.iterrows()])
    ws.cell(row=r+1, column=7, value="Freight Charge").font = b_bold; ws.cell(row=r+1, column=8, value=fr).border = tb
    ws.cell(row=r+2, column=7, value="GST Total").font = b_bold; ws.cell(row=r+2, column=8, value=ti - sub).border = tb
    ws.cell(row=r+3, column=7, value="Total Order Value").font = b_bold; ws.cell(row=r+3, column=8, value=ti + fr).border = tb
    
    ws.column_dimensions['A'].width = 15; ws.column_dimensions['B'].width = 15; ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40; ws.column_dimensions['F'].width = 15; ws.column_dimensions['H'].width = 15; ws.column_dimensions['G'].width = 15
    
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── PO PDF Export ───────────────────────────────────────────
def export_po_pdf(quote_no: str, df: pd.DataFrame) -> io.BytesIO:
    try:
        from fpdf import FPDF
    except ImportError:
        return None

    q_df = df[df["Quote No"] == quote_no]
    if q_df.empty:
        return None
    row1 = q_df.iloc[0]
    loc_code = str(row1.get("Location Code", ""))
    loc_info = LOCATION_DETAILS.get(loc_code, {})

    class PDF(FPDF):
        def header(self):
            logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
            if os.path.exists(logo_path):
                self.image(logo_path, 10, 8, 45)
            self.set_font("helvetica", "B", 14)
            self.set_xy(60, 10)
            self.cell(0, 8, "SUPREME COMPUTERS INDIA PVT. LTD.", 0, 1, "C")
            self.set_font("helvetica", "", 9)
            self.set_x(60)
            self.cell(0, 5, "18/18, Majestic Plaza, Narasingapuram Street, Mount Road, Chennai - 600002", 0, 1, "C")
            gstin_val = loc_info.get("GSTIN", "33AAOCS1408H1ZR")
            self.set_x(60)
            self.cell(0, 5, f"GSTIN: {gstin_val}  |  Tel: 044-42082020", 0, 1, "C")
            self.ln(4)
            self.set_fill_color(31, 73, 125)
            self.set_text_color(255, 255, 255)
            self.set_font("helvetica", "B", 12)
            self.cell(0, 8, "PURCHASE ORDER", 0, 1, "C", True)
            self.set_text_color(0, 0, 0)
            self.ln(3)

    pdf = PDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font("helvetica", "", 9)

    # PO & Vendor Details
    vendor_name = str(row1.get("Vendor Name", ""))
    v_addr = str(row1.get("Vendor Address", ""))
    v_city = str(row1.get("Vendor City", ""))
    v_gst = str(row1.get("Vendor GST", ""))
    v_email = str(row1.get("Vendor Email", ""))
    
    if (not v_addr or v_addr == "nan") and vendor_name:
        vendor_df = load_master_sheet("Vendor Master")
        if not vendor_df.empty:
            vrow = vendor_df[vendor_df["Name"] == vendor_name] if "Name" in vendor_df.columns else vendor_df[vendor_df["name"] == vendor_name]
            if not vrow.empty:
                vr = vrow.iloc[0]
                v_addr = str(vr.get("address", vr.get("Address", "")))
                v_city = str(vr.get("city", vr.get("City", "")))
                v_gst = str(vr.get("vatRegistrationNo", vr.get("GST", vr.get("GSTIN", ""))))
                v_email = str(vr.get("email", vr.get("Email", vr.get("E-Mail", ""))))
                
    v_addr = v_addr if v_addr != "nan" else ""
    v_city = v_city if v_city != "nan" else ""
    v_gst = v_gst if v_gst != "nan" else ""
    v_email = v_email if v_email != "nan" else ""
    
    pdf.set_fill_color(189, 215, 238)
    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "PO Number:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(60, 7, str(quote_no), 1, 0, "L")
    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "PO Date:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(60, 7, str(row1.get("Quote Date", "")), 1, 1, "L")

    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "Vendor Name:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(155, 7, vendor_name, 1, 1, "L")

    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "Vendor Address:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    full_addr = v_addr
    if v_city: full_addr += f", {v_city}"
    pdf.cell(155, 7, full_addr, 1, 1, "L")

    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "Vendor GSTIN:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(60, 7, v_gst, 1, 0, "L")
    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "Vendor Email:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(60, 7, v_email, 1, 1, "L")

    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "Vendor No:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(60, 7, str(row1.get("Vendor No", "")), 1, 0, "L")
    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "Payment Terms:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(60, 7, str(row1.get("Credit Days", "")), 1, 1, "L")

    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "Location:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(60, 7, loc_code, 1, 0, "L")
    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "Payment Method:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(60, 7, str(row1.get("Payment Method", "")), 1, 1, "L")
    
    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "Mode of Delivery:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(60, 7, str(row1.get("Mode of Delivery", "")), 1, 0, "L")
    pdf.set_font("helvetica", "B", 9)
    pdf.cell(35, 7, "Delivery Date:", 1, 0, "L", True)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(60, 7, str(row1.get("Expected Delivery Date", "")), 1, 1, "L")
    
    c_details = str(row1.get("Courier Details", ""))
    if c_details:
        pdf.set_font("helvetica", "B", 9)
        pdf.cell(35, 7, "Courier/Transport:", 1, 0, "L", True)
        pdf.set_font("helvetica", "", 9)
        pdf.cell(155, 7, c_details, 1, 1, "L")
        
    pdf.ln(3)

    # Delivery Address
    pdf.set_fill_color(31, 73, 125)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", "B", 9)
    pdf.cell(0, 7, "Delivery Address", 1, 1, "L", True)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("helvetica", "", 9)
    addr_text = loc_info.get("Address", "")
    contact_text = loc_info.get("Contact", "")
    pdf.multi_cell(0, 6, f"{addr_text}\nContact: {contact_text}", 1)
    pdf.ln(3)

    # Line Items Table Header
    pdf.set_fill_color(192, 0, 0)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", "B", 8)
    col_w = [10, 25, 25, 60, 10, 22, 10, 28] # Adjusted widths
    headers = ["S.No", "ERP Code", "Vendor Item", "Description", "Qty", "Price(Excl.)", "GST%", "Total(Incl.)"]
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, 1, 0, "C", True)
    pdf.ln()

    # Rows
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("helvetica", "", 8)
    grand_total = 0.0
    sub_total = 0.0
    for _, l in q_df.iterrows():
        qty = float(l.get("Qty", 0) or 0)
        pb = float(l.get("Price Before GST", 0) or 0)
        pi = float(l.get("Price Inc. GST", 0) or 0)
        row_total = qty * pi
        grand_total += row_total
        sub_total += qty * pb
        desc = str(l.get("Product Description", ""))[:38]
        erp = str(l.get("ERP Code", ""))[:14]
        vin = str(l.get("Vendor Item No", ""))[:14]
        pdf.cell(col_w[0], 6, str(l.get("S.No", "")), 1, 0, "C")
        pdf.cell(col_w[1], 6, erp, 1, 0, "L")
        pdf.cell(col_w[2], 6, vin, 1, 0, "L")
        pdf.cell(col_w[3], 6, desc, 1, 0, "L")
        pdf.cell(col_w[4], 6, str(int(qty)), 1, 0, "C")
        pdf.cell(col_w[5], 6, f"{pb:,.2f}", 1, 0, "R")
        pdf.cell(col_w[6], 6, str(l.get("GST %", "")), 1, 0, "C")
        pdf.cell(col_w[7], 6, f"{row_total:,.2f}", 1, 1, "R")

    # Totals
    fr = float(row1.get("Freight Charge", 0) or 0)
    gst_total = grand_total - sub_total
    total_cols_w = sum(col_w[:-1])
    pdf.set_font("helvetica", "B", 9)
    pdf.cell(total_cols_w, 7, "Sub Total (Excl. GST):", 0, 0, "R")
    pdf.cell(col_w[-1], 7, f"{sub_total:,.2f}", 1, 1, "R")
    pdf.cell(total_cols_w, 7, "Total GST:", 0, 0, "R")
    pdf.cell(col_w[-1], 7, f"{gst_total:,.2f}", 1, 1, "R")
    pdf.cell(total_cols_w, 7, "Freight Charge:", 0, 0, "R")
    pdf.cell(col_w[-1], 7, f"{fr:,.2f}", 1, 1, "R")

    pdf.set_fill_color(31, 73, 125)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(total_cols_w, 8, "GRAND TOTAL (Incl. GST):  ", 0, 0, "R", True)
    pdf.cell(col_w[-1], 8, f"{grand_total + fr:,.2f}", 1, 1, "R", True)
    pdf.set_text_color(0, 0, 0)

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf

# ── DASHBOARD ───────────────────────────────────────────────
def page_dashboard():
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use("Agg")

    # Full View Logo at the top
    try:
        st.image("logo.png", use_container_width=True)
    except:
        pass

    # ── Premium Header Bar ──
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 100%); padding: 18px 28px; border-radius: 14px; margin-bottom: 18px;">
        <span style="color: #ffffff; font-size: 1.5rem; font-weight: 700; letter-spacing: 0.5px;">📊 Purchase Quote Executive Dashboard</span>
    </div>
    """, unsafe_allow_html=True)
    df = load_quotes()

    if df.empty:
        st.info("No quotes found. Create your first quote!")
        return

    # Prepare data
    df_temp = df.copy()
    df_temp["Qty_num"] = pd.to_numeric(df_temp["Qty"], errors="coerce").fillna(0.0)
    df_temp["Price_Inc_num"] = pd.to_numeric(df_temp["Price Inc. GST"], errors="coerce").fillna(0.0)
    df_temp["Line_Total_Inc"] = df_temp["Qty_num"] * df_temp["Price_Inc_num"]
    df_temp["Created_DT"] = pd.to_datetime(df_temp["Created Date"], errors="coerce")

    # ── Filter Row ──
    fc1, fc2, fc3, fc4 = st.columns(4)
    with fc1:
        date_opt = st.selectbox("📅 Time Period", ["All Time", "Today", "This Week", "This Month", "This Year", "Custom Range"])
    with fc2:
        unique_vendors = ["All Vendors"] + sorted(df["Vendor Name"].dropna().unique().tolist())
        sel_vendor = st.selectbox("🏢 Vendor", unique_vendors)
    with fc3:
        unique_locs = ["All Locations"] + sorted(df["Location Code"].dropna().unique().tolist())
        sel_loc = st.selectbox("📍 Location", unique_locs)
    with fc4:
        unique_purchasers = ["All Purchasers"] + sorted(df["Purchase Quote raised by"].dropna().unique().tolist())
        sel_purchaser = st.selectbox("👤 Purchaser", unique_purchasers)

    start_date, end_date = None, None
    if date_opt == "Custom Range":
        custom_range = st.date_input("Select Date Range", value=(date.today(), date.today()))
        if isinstance(custom_range, tuple) and len(custom_range) == 2:
            start_date, end_date = custom_range
    elif date_opt == "Today":
        start_date = end_date = date.today()
    elif date_opt == "This Week":
        start_date = date.today() - pd.Timedelta(days=date.today().weekday()); end_date = date.today()
    elif date_opt == "This Month":
        start_date = date(date.today().year, date.today().month, 1); end_date = date.today()
    elif date_opt == "This Year":
        start_date = date(date.today().year, 1, 1); end_date = date.today()

    # Apply all filters
    fdf = df_temp.copy()
    if start_date and end_date:
        fdf = fdf[(fdf["Created_DT"] >= pd.to_datetime(start_date)) & (fdf["Created_DT"] < pd.to_datetime(end_date) + pd.Timedelta(days=1))]
    if sel_vendor != "All Vendors":
        fdf = fdf[fdf["Vendor Name"] == sel_vendor]
    if sel_loc != "All Locations":
        fdf = fdf[fdf["Location Code"] == sel_loc]
    if sel_purchaser != "All Purchasers":
        fdf = fdf[fdf["Purchase Quote raised by"] == sel_purchaser]

    # ── Aggregations ──
    total_quotes = fdf["Quote No"].nunique() if not fdf.empty else 0
    grand_qty = fdf["Qty_num"].sum() if not fdf.empty else 0
    total_locs = fdf["Location Code"].nunique() if not fdf.empty else 0

    if not fdf.empty:
        qt = fdf.groupby("Quote No").agg(
            line_total=("Line_Total_Inc", "sum"),
            freight=("Freight Charge", lambda x: pd.to_numeric(x.iloc[0], errors="coerce") if not x.empty else 0.0),
            qty=("Qty_num", "sum"), purchaser=("Purchase Quote raised by", "first"),
            vendor=("Vendor Name", "first"), location=("Location Code", "first"),
            status=("Status", "first"), created=("Created_DT", "first")
        ).reset_index()
        qt["freight"] = qt["freight"].fillna(0.0)
        qt["value"] = qt["line_total"] + qt["freight"]
        grand_value = qt["value"].sum()
    else:
        qt = pd.DataFrame(columns=["Quote No","value","qty","purchaser","vendor","location","status","created"])
        grand_value = 0.0

    # ── Navy color palette for matplotlib ──
    NAVY = "#0f172a"; NAVY2 = "#1e3a5f"; NAVY3 = "#2d5a8e"; LIGHT = "#e2e8f0"
    PIE_COLORS = ["#0f172a", "#1e3a5f", "#2d5a8e", "#4a90c4", "#7cb3d4", "#a8d0e6"]

    # ── 4 KPI Cards ──
    st.markdown("")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("📊 Total Quotes Raised", f"{total_quotes}")
    k2.metric("📦 Total Qty Purchased", f"{int(grand_qty):,}")
    k3.metric("💰 Total Value (₹)", f"{grand_value:,.2f}")
    k4.metric("📍 Total Locations", f"{total_locs}")

    st.markdown("---")

    # ── Row 2: Breakdown Table (Full Width) ──
    st.markdown("#### 📋 Purchaser | Vendor | Location Breakdown")
    if not qt.empty:
        bd = qt.groupby(["purchaser", "vendor", "location"]).agg(
            quotes=("Quote No", "nunique"), total_val=("value", "sum")
        ).reset_index()
        bd.columns = ["Purchased By", "Vendor", "Location", "Total Quotes Raised", "Total Value (₹)"]
        bd = bd.sort_values("Total Value (₹)", ascending=False)
        bd_disp = bd.copy()
        bd_disp["Total Value (₹)"] = bd_disp["Total Value (₹)"].apply(lambda x: f"₹{x:,.2f}")
        st.dataframe(bd_disp, use_container_width=True, hide_index=True)
    else:
        st.info("No data for selected filters.")

    st.markdown("---")

    # ── Row 3: 4 Charts ──
    ch1, ch2, ch3, ch4 = st.columns(4)

    with ch1:
        st.markdown("#### 💹 Value by Vendor")
        if not qt.empty:
            vv = qt.groupby("vendor")["value"].sum().sort_values(ascending=False).head(6)
            fig2, ax2 = plt.subplots(figsize=(4, 3.5))
            fig2.patch.set_facecolor("white")
            bars = ax2.bar(range(len(vv)), vv.values, color=[NAVY, NAVY2, NAVY3, "#4a90c4", "#7cb3d4", "#a8d0e6"][:len(vv)])
            ax2.set_xticks(range(len(vv)))
            ax2.set_xticklabels([v[:12] for v in vv.index], rotation=45, ha="right", fontsize=7, color=NAVY)
            ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x/1000:.0f}K" if x >= 1000 else f"{x:.0f}"))
            ax2.tick_params(axis="y", labelsize=8, colors=NAVY); ax2.spines[["top","right"]].set_visible(False)
            plt.tight_layout()
            st.pyplot(fig2, use_container_width=True); plt.close(fig2)

    with ch2:
        st.markdown("#### 📍 Quotes by Location")
        if not qt.empty:
            ql = qt.groupby("location")["Quote No"].nunique().sort_values(ascending=False).head(6)
            fig3, ax3 = plt.subplots(figsize=(4, 3.5))
            fig3.patch.set_facecolor("white")
            ax3.bar(range(len(ql)), ql.values, color=[NAVY, NAVY2, NAVY3, "#4a90c4"][:len(ql)])
            ax3.set_xticks(range(len(ql)))
            ax3.set_xticklabels([l[:10] for l in ql.index], rotation=45, ha="right", fontsize=7, color=NAVY)
            ax3.tick_params(axis="y", labelsize=8, colors=NAVY); ax3.spines[["top","right"]].set_visible(False)
            plt.tight_layout()
            st.pyplot(fig3, use_container_width=True); plt.close(fig3)

    with ch3:
        st.markdown("#### 👤 Qty by Purchaser")
        if not qt.empty:
            qp = qt.groupby("purchaser")["qty"].sum().sort_values(ascending=True)
            fig4, ax4 = plt.subplots(figsize=(4, 3.5))
            fig4.patch.set_facecolor("white")
            ax4.barh(range(len(qp)), qp.values, color=[NAVY, NAVY2, NAVY3, "#4a90c4"][:len(qp)])
            ax4.set_yticks(range(len(qp)))
            ax4.set_yticklabels([p[:12] for p in qp.index], fontsize=8, color=NAVY)
            ax4.tick_params(axis="x", labelsize=8, colors=NAVY); ax4.spines[["top","right"]].set_visible(False)
            plt.tight_layout()
            st.pyplot(fig4, use_container_width=True); plt.close(fig4)

    with ch4:
        st.markdown("#### 📈 Quotes Over Time")
        if not qt.empty and qt["created"].notna().any():
            qt_time = qt.copy()
            qt_time["month"] = qt_time["created"].dt.to_period("M").astype(str)
            monthly = qt_time.groupby("month")["Quote No"].nunique().reset_index()
            monthly.columns = ["Month", "Quotes"]
            fig5, ax5 = plt.subplots(figsize=(4, 3.5))
            fig5.patch.set_facecolor("white")
            ax5.plot(monthly["Month"], monthly["Quotes"], color=NAVY, marker="o", linewidth=2, markersize=5)
            ax5.fill_between(monthly["Month"], monthly["Quotes"], alpha=0.1, color=NAVY2)
            ax5.set_xticklabels(monthly["Month"], rotation=45, ha="right", fontsize=7, color=NAVY)
            ax5.tick_params(axis="y", labelsize=8, colors=NAVY); ax5.spines[["top","right"]].set_visible(False)
            plt.tight_layout()
            st.pyplot(fig5, use_container_width=True); plt.close(fig5)

    st.markdown("---")

    # ── Row 4: Monthly Trend + Value Contribution Pie ──
    bot1, bot2 = st.columns([3, 2])

    with bot1:
        st.markdown("#### 📉 Monthly Value Trend (₹)")
        if not qt.empty and qt["created"].notna().any():
            qt_t2 = qt.copy()
            qt_t2["month"] = qt_t2["created"].dt.to_period("M").astype(str)
            mv = qt_t2.groupby("month")["value"].sum().reset_index()
            mv.columns = ["Month", "Value"]
            fig6, ax6 = plt.subplots(figsize=(8, 3.5))
            fig6.patch.set_facecolor("white")
            ax6.plot(mv["Month"], mv["Value"], color=NAVY, marker="o", linewidth=2.5, markersize=6)
            ax6.fill_between(mv["Month"], mv["Value"], alpha=0.08, color=NAVY2)
            ax6.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"₹{x/1000:.0f}K" if x >= 1000 else f"₹{x:.0f}"))
            ax6.set_xticklabels(mv["Month"], rotation=45, ha="right", fontsize=8, color=NAVY)
            ax6.tick_params(axis="y", labelsize=8, colors=NAVY); ax6.spines[["top","right"]].set_visible(False)
            plt.tight_layout()
            st.pyplot(fig6, use_container_width=True); plt.close(fig6)

    with bot2:
        st.markdown("#### 🥧 Value Contribution (₹)")
        if not qt.empty:
            vc = qt.groupby("vendor")["value"].sum().sort_values(ascending=False).head(5)
            fig7, ax7 = plt.subplots(figsize=(4, 3.5))
            fig7.patch.set_facecolor("white")
            wedges, texts = ax7.pie(vc.values, labels=[v[:14] for v in vc.index],
                colors=PIE_COLORS[:len(vc)], textprops={"fontsize": 8, "color": NAVY}, startangle=90)
            ax7.set_aspect("equal")
            st.pyplot(fig7, use_container_width=True); plt.close(fig7)

    st.markdown("---")

    # ── Recent Quotes Table ──
    st.markdown("#### 📄 Recent Purchase Quotes")
    if not fdf.empty:
        recent = fdf.drop_duplicates("Quote No", keep="last").sort_values("Created Date", ascending=False).head(10)
        if not qt.empty:
            recent = recent.merge(qt[["Quote No", "value"]], on="Quote No", how="left")
            recent = recent.rename(columns={"value": "Total Order Value"})
        disp_r = recent[["Quote No", "Purchase Quote raised by", "Vendor Name", "Location Code", "Status", "Total Order Value", "Created Date"]].copy()
        disp_r["Total Order Value"] = disp_r["Total Order Value"].apply(lambda x: f"₹{x:,.2f}" if pd.notnull(x) else "₹0.00")
        st.dataframe(disp_r, use_container_width=True, hide_index=True)
    else:
        st.info("No quotes for selected filters.")

# ── CREATE / EDIT QUOTE ──────────────────────────────────────────
def page_create_quote():
    st.markdown('<div class="section-header">➕ Create / Edit Purchase Quote</div>', unsafe_allow_html=True)

    quotes_df = load_quotes()
    
    # ── Edit Mode Selection ──
    edit_mode = False
    sel_quote_to_edit = "-- New Quote --"
    if not quotes_df.empty:
        existing_quotes = sorted(quotes_df["Quote No"].dropna().unique().tolist(), reverse=True)
        sel_quote_to_edit = st.selectbox("📝 Create New or Edit Existing Quote", ["-- New Quote --"] + existing_quotes)
        if sel_quote_to_edit != "-- New Quote --":
            edit_mode = True

    # ── Logic to load existing data if in edit mode ──
    # Reset submit guard whenever a different quote is loaded
    if edit_mode and st.session_state.get("last_loaded_quote") != sel_quote_to_edit:
        st.session_state.quote_submitted  = False
        st.session_state.submit_validated = False
    if edit_mode and ("last_loaded_quote" not in st.session_state or st.session_state.last_loaded_quote != sel_quote_to_edit):
        q_data = quotes_df[quotes_df["Quote No"] == sel_quote_to_edit]
        if not q_data.empty:
            first_row = q_data.iloc[0]
            st.session_state.edit_salesperson = str(first_row.get("Purchase Quote raised by", ""))
            st.session_state.edit_vendor = str(first_row.get("Vendor Name", ""))
            st.session_state.edit_location = str(first_row.get("Location Code", "EGMR"))
            st.session_state.edit_status = str(first_row.get("Status", "Pending Approval"))
            st.session_state.edit_payment_method = str(first_row.get("Payment Method", ""))
            st.session_state.edit_credit_days = str(first_row.get("Credit Days", ""))
            st.session_state.edit_mode_of_delivery = str(first_row.get("Mode of Delivery", ""))
            st.session_state.edit_courier_details = str(first_row.get("Courier Details", ""))
            
            # Load Expected Delivery Date
            ed_val = first_row.get("Expected Delivery Date", None)
            try:
                if pd.notnull(ed_val) and str(ed_val).strip():
                    st.session_state.edit_expected_delivery = pd.to_datetime(ed_val).date()
                else:
                    st.session_state.edit_expected_delivery = None
            except:
                st.session_state.edit_expected_delivery = None
            
            # Load line items
            new_lines = []
            for _, row in q_data.iterrows():
                erp_v = str(row.get("ERP Code", "")).strip()
                is_new_v = (erp_v == "NEW")
                try:
                    qty_v = int(float(row.get("Qty", 1) or 1))
                except (ValueError, TypeError):
                    qty_v = 1
                try:
                    price_v = float(row.get("Price Before GST", 0.0) or 0.0)
                except (ValueError, TypeError):
                    price_v = 0.0
                try:
                    gst_v = float(row.get("GST %", 18.0) or 18.0)
                except (ValueError, TypeError):
                    gst_v = 18.0
                new_lines.append({
                    "erp_code":         erp_v,
                    "vendor_item_no":   str(row.get("Vendor Item No", "")),
                    "description":      str(row.get("Product Description", "")),
                    "qty":              qty_v,
                    "price_before_gst": price_v,
                    "gst_percent":      gst_v,
                    "remarks":          str(row.get("Remarks", "")),
                    "is_new_item":      is_new_v,
                })
            st.session_state.line_items = new_lines
            st.session_state.last_loaded_quote = sel_quote_to_edit
    elif not edit_mode:
        if "last_loaded_quote" in st.session_state:
            del st.session_state.last_loaded_quote
            st.session_state.line_items = []
            st.session_state.quote_submitted  = False
            st.session_state.submit_validated = False
            if "submitted_quote_no" in st.session_state:
                del st.session_state.submitted_quote_no

    quote_no = sel_quote_to_edit if edit_mode else next_quote_number(quotes_df)

    # ── Header section ───────────────────────────────────────
    st.markdown("---")
    st.subheader("Quote Header")
    h1, h2, h3, h4 = st.columns(4)
    with h1:
        st.text_input("Quote Number", value=quote_no, disabled=True)
    with h2:
        quote_date = st.date_input("Quote Date", value=date.today())
    with h3:
        location_opts = sorted(list(LOCATION_DETAILS.keys()))
        def_loc = st.session_state.get("edit_location", "EGMR")
        loc_idx = location_opts.index(def_loc) if def_loc in location_opts else 0
        location_code = st.selectbox("Location Code", location_opts, index=loc_idx)
        
        # Display selected location details briefly
        loc_info = LOCATION_DETAILS.get(location_code, {})
        if loc_info:
            st.caption(f"📍 {loc_info['Address'][:50]}...")
    with h4:
        salesperson = st.text_input("Purchase Quote raised by", value=st.session_state.get("edit_salesperson", ""), placeholder="Enter your name")

    # Vendor selection
    st.markdown("#### Vendor")
    vendor_names = []
    if not vendor_df.empty:
        for nc in ["name", "Name", "searchName", "SearchName"]:
            if nc in vendor_df.columns:
                vendor_names = vendor_df[nc].dropna().unique().tolist()
                break

    v1, v2, v3, v4 = st.columns([2.5, 1, 3.5, 1])
    with v1:
        def_v = st.session_state.get("edit_vendor", "-- Select Vendor --")
        v_idx = vendor_names.index(def_v) + 1 if def_v in vendor_names else 0
        selected_vendor = st.selectbox("Vendor Name", ["-- Select Vendor --"] + vendor_names, index=v_idx)
    
    vendor_no = ""
    vendor_addr = ""
    vendor_city = ""
    vendor_gst = ""
    vendor_email = ""
    v_pay_meth = ""
    v_cred_days = ""
    
    if selected_vendor != "-- Select Vendor --" and not vendor_df.empty:
        vrow = vendor_df[vendor_df["Name"] == selected_vendor] if "Name" in vendor_df.columns else vendor_df[vendor_df["name"] == selected_vendor]
        if not vrow.empty:
            vr = vrow.iloc[0]
            vendor_no = str(vr.get("number", vr.get("vendorNo", "")))
            vendor_addr = str(vr.get("address", vr.get("Address", "")))
            vendor_city = str(vr.get("city", vr.get("City", "")))
            vendor_gst = str(vr.get("vatRegistrationNo", vr.get("GST", vr.get("GSTIN", ""))))
            vendor_email = str(vr.get("email", vr.get("Email", vr.get("E-Mail", ""))))
            v_pay_meth = str(vr.get("paymentMethodCode", ""))
            v_cred_days = str(vr.get("paymentTermsCode", ""))

    with v2:
        st.text_input("Vendor No", value=vendor_no, disabled=True)
    with v3:
        st.text_input("Vendor Address", value=vendor_addr, disabled=True)
    with v4:
        st.text_input("City", value=vendor_city, disabled=True)
    
    if selected_vendor != "-- Select Vendor --":
        details = []
        details.append(f"🏢 **Name:** {selected_vendor}")
        if vendor_addr: details.append(f"📍 **Address:** {vendor_addr}")
        if vendor_city: details.append(f"🏙️ **City:** {vendor_city}")
        if vendor_gst: details.append(f"🛡️ **GST:** {vendor_gst}")
        st.caption("  •  ".join(details))

    # ── Payment & Delivery ───────────────────────────────────
    st.markdown("---")
    st.subheader("Payment & Delivery Options")
    
    p1, p2, p3 = st.columns(3)
    pm_opts = ["", "Cash", "Online Transfer", "CDC", "PDC"]
    
    # Use existing data if editing, otherwise vendor default
    if edit_mode:
        curr_pm = st.session_state.get("edit_payment_method", "")
        pm_idx = pm_opts.index(curr_pm) if curr_pm in pm_opts else 0
    else:
        pm_idx = pm_opts.index(v_pay_meth) if v_pay_meth in pm_opts else 0
        
    with p1: 
        payment_method = st.selectbox("Payment Method", pm_opts, index=pm_idx)
    with p2: 
        credit_days = st.text_input("Credit Days", value=st.session_state.get("edit_credit_days", v_cred_days))
    
    md_opts = ["", "Door Delivery", "Pickup", "Courier"]
    def_md = st.session_state.get("edit_mode_of_delivery", "")
    md_idx = md_opts.index(def_md) if def_md in md_opts else 0
    with p3: 
        mode_of_delivery = st.selectbox("Mode of Delivery", md_opts, index=md_idx)
        
    d1, d2 = st.columns(2)
    with d1: 
        def_ed = st.session_state.get("edit_expected_delivery", None)
        expected_delivery = st.date_input("Expected Delivery Date", value=def_ed)
    with d2:
        courier_details = st.text_input("Courier / Transport Details", value=st.session_state.get("edit_courier_details", ""))
        
    cash = "Yes" if payment_method == "Cash" else ""
    online_transfer = "Yes" if payment_method == "Online Transfer" else ""
    cdc = "Yes" if payment_method == "CDC" else ""
    pdc = "Yes" if payment_method == "PDC" else ""
    
    door_delivery = "Yes" if mode_of_delivery == "Door Delivery" else ""
    pickup = "Yes" if mode_of_delivery == "Pickup" else ""
    courier = "Yes" if mode_of_delivery == "Courier" else ""

    # ── Line Items ───────────────────────────────────────────
    st.markdown("---")
    st.subheader("Line Items")

    # ── Build master lookup (ERP / VendorItem / Description searchable) ──
    _SEARCH_SENTINEL = "🔍 Type to search or ➕ enter new item below..."
    _NEW_SENTINEL    = "✏️ NEW ITEM (manual entry)"

    item_lookup = {}        # display_label -> field dict
    search_options = [_SEARCH_SENTINEL, _NEW_SENTINEL]

    if not item_df.empty:
        _erp_col   = next((c for c in ["No", "Item_No", "ERP_Code"] if c in item_df.columns), None)
        _desc_col  = next((c for c in ["Description", "Search_Description", "Product Description",
                                       "Product_Description", "Item Description", "Name"] if c in item_df.columns), None)
        _vin_col   = next((c for c in ["Vendor_Item_No", "VendorItemNo", "Vendor Item No", "Vendor_No"] if c in item_df.columns), None)
        _price_col = next((c for c in ["Quoting_Price_WIN", "Unit_Price", "Last_Direct_Cost"] if c in item_df.columns), None)
        _gst_col   = next((c for c in ["GST", "GST%", "GST_Percent", "Tax_Group_Code"] if c in item_df.columns), None)

        for _, irow in item_df.iterrows():
            erp_v  = str(irow[_erp_col]).strip()  if _erp_col  else ""
            desc_v = str(irow[_desc_col]).strip() if _desc_col else ""
            vin_v  = str(irow[_vin_col]).strip()  if _vin_col  else ""
            try:
                price_v = float(irow[_price_col]) if _price_col else 0.0
            except (ValueError, TypeError):
                price_v = 0.0
            try:
                gst_v = float(irow[_gst_col]) if _gst_col else 18.0
            except (ValueError, TypeError):
                gst_v = 18.0
            if not erp_v or erp_v.lower() == "nan":
                continue
            price_str = f"₹{price_v:,.2f}" if price_v else ""
            label = " | ".join(filter(None, [erp_v, vin_v, desc_v, price_str]))
            item_lookup[label] = {
                "erp_code": erp_v, "vendor_item_no": vin_v,
                "description": desc_v, "price_before_gst": price_v, "gst_percent": gst_v,
            }
            search_options.append(label)

    def _label_for_erp(erp_code):
        for lbl, data in item_lookup.items():
            if data["erp_code"] == erp_code:
                return lbl
        return _SEARCH_SENTINEL

    if "line_items" not in st.session_state:
        st.session_state.line_items = []

    # Reset submit guard whenever the line items section is rebuilt fresh
    # (only reset it if we just switched to New Quote mode)
    if not edit_mode and not st.session_state.get("submitted_quote_no"):
        st.session_state.quote_submitted  = False
        st.session_state.submit_validated = False

    # ── Toolbar ──────────────────────────────────────────────
    col_add, col_clear, col_hint = st.columns([1, 1.4, 5])
    with col_add:
        if st.button("➕ Add Line", use_container_width=True):
            st.session_state.line_items.append({
                "erp_code":         "",
                "vendor_item_no":   "",
                "description":      "",
                "qty":              None,    # blank until user types
                "price_before_gst": None,   # blank until user types
                "gst_percent":      18.0,
                "remarks":          "",
                "is_new_item":      False,
            })
            st.session_state.quote_submitted  = False
            st.session_state.submit_validated = False
    with col_clear:
        if st.button("🗑️ Clear All Lines", use_container_width=True):
            st.session_state.line_items = []
            st.session_state.quote_submitted  = False
            st.session_state.submit_validated = False
            if "submitted_quote_no" in st.session_state:
                del st.session_state.submitted_quote_no
    with col_hint:
        st.caption(
            "💡 Search by ERP Code / Vendor Item No / Description  •  "
            "Choose **'NEW ITEM'** to enter an unlisted item manually  •  "
            "Price is always editable (PO-level override, master is never changed)"
        )

    line_data_for_saving = []
    items_to_remove      = []
    line_errors          = []   # populated only after submit attempt

    # ── Column header strip ───────────────────────────────────
    st.markdown(
        "<div class='li-header'>"
        "<span style='width:2%'>#</span>"
        "<span style='width:22%'>Item Search</span>"
        "<span style='width:13%'>Vendor Item No</span>"
        "<span style='width:23%'>Description</span>"
        "<span style='width:7%'>Qty</span>"
        "<span style='width:12%'>Price (Excl. GST)</span>"
        "<span style='width:7%'>GST %</span>"
        "<span style='width:11%'>Total (Incl. GST)</span>"
        "<span style='width:3%'></span>"
        "</div>",
        unsafe_allow_html=True,
    )

    for i, line in enumerate(st.session_state.line_items):
        is_new = line.get("is_new_item", False)

        # Widths: #(tiny) | search(wide) | vin | desc(wide) | qty | price | gst | total | del
        c0, c1, c2, c3, c4, c5, c6, c7, c8 = st.columns([0.4, 3.2, 2.0, 3.0, 1.0, 1.8, 1.0, 1.8, 0.4])

        with c0:
            st.markdown(
                f"<div style='padding-top:32px;font-weight:700;font-size:13px;"
                f"color:#94a3b8;text-align:center'>{i+1}</div>",
                unsafe_allow_html=True,
            )

        # ── Item selector ─────────────────────────────────────
        with c1:
            if is_new:
                st.markdown(
                    "<span style='font-size:11px;color:#f59e0b;font-weight:700;"
                    "background:#422006;padding:2px 8px;border-radius:4px'>✏️ NEW ITEM</span>",
                    unsafe_allow_html=True,
                )
                if st.button("↩ Select from list", key=f"back_list_{i}", use_container_width=True):
                    st.session_state.line_items[i]["is_new_item"] = False
                    st.session_state.line_items[i]["erp_code"]    = ""
                    st.rerun()
            else:
                current_label = _label_for_erp(line["erp_code"]) if line["erp_code"] else _SEARCH_SENTINEL
                try:
                    sel_idx = search_options.index(current_label)
                except ValueError:
                    sel_idx = 0
                chosen_label = st.selectbox(
                    f"Item {i+1}", search_options, index=sel_idx,
                    key=f"item_sel_{i}", label_visibility="collapsed",
                )
                if chosen_label == _NEW_SENTINEL:
                    st.session_state.line_items[i].update({
                        "is_new_item": True, "erp_code": "NEW",
                        "vendor_item_no": "", "description": "",
                        "qty": None, "price_before_gst": None, "gst_percent": 18.0,
                    })
                    st.rerun()
                elif chosen_label in item_lookup:
                    if chosen_label != _label_for_erp(line["erp_code"]):
                        data = item_lookup[chosen_label]
                        # Auto-fill from master; qty stays blank (user must enter)
                        st.session_state.line_items[i].update({
                            "erp_code":         data["erp_code"],
                            "vendor_item_no":   data["vendor_item_no"],
                            "description":      data["description"],
                            "price_before_gst": data["price_before_gst"] if data["price_before_gst"] else None,
                            "gst_percent":      data["gst_percent"],
                            "is_new_item":      False,
                            "qty":              None,   # always blank on item change
                        })
                        st.rerun()

        # ── Refresh live values after possible rerun triggers ──
        live   = st.session_state.line_items[i]
        is_new = live.get("is_new_item", False)

        # ── Is this row "touched" at all? ─────────────────────
        # A row is considered blank if: no item selected, no VIN typed,
        # qty is None/0, price is None/0.  Blank rows are silently skipped.
        row_erp  = live.get("erp_code", "")
        row_vin_stored  = live.get("vendor_item_no", "")
        row_touched = bool(row_erp and row_erp not in ("", _SEARCH_SENTINEL))

        # ── Vendor Item No ────────────────────────────────────
        with c2:
            if is_new:
                vin_input = st.text_input(
                    f"VIN {i}", value=live.get("vendor_item_no", ""),
                    placeholder="Vendor Item No *",
                    key=f"vin_{i}", label_visibility="collapsed",
                )
                # A new-item row is touched as soon as VIN is typed
                if vin_input.strip():
                    row_touched = True
            else:
                vin_input = row_vin_stored
                st.text_input(
                    f"VIN {i}", value=vin_input,
                    disabled=True, label_visibility="collapsed",
                )

        # ── Description ───────────────────────────────────────
        with c3:
            if is_new:
                desc_input = st.text_input(
                    f"Desc {i}", value=live.get("description", ""),
                    placeholder="Product Description",
                    key=f"desc_{i}", label_visibility="collapsed",
                )
            else:
                desc_input = live.get("description", "")
                st.text_input(
                    f"Desc {i}", value=desc_input,
                    disabled=True, label_visibility="collapsed",
                )

        # ── Qty — blank by default (value=None rendered as empty) ─
        with c4:
            stored_qty = live.get("qty")
            qty = st.number_input(
                f"Qty {i}",
                min_value=0,
                value=int(stored_qty) if stored_qty is not None else None,
                step=1,
                key=f"qty_{i}",
                label_visibility="collapsed",
                placeholder="Qty",
            )
            if qty and qty > 0:
                row_touched = True

        # ── Price — blank by default ──────────────────────────
        with c5:
            stored_price = live.get("price_before_gst")
            price = st.number_input(
                f"Price {i}",
                min_value=0.0,
                value=float(stored_price) if stored_price is not None else None,
                step=0.01,
                format="%.2f",
                key=f"price_{i}",
                label_visibility="collapsed",
                placeholder="0.00",
            )
            if price and price > 0:
                row_touched = True

        # ── GST % ─────────────────────────────────────────────
        with c6:
            gst_p = st.number_input(
                f"GST% {i}",
                min_value=0.0,
                max_value=100.0,
                value=float(live.get("gst_percent") or 18.0),
                step=0.5,
                format="%.1f",
                key=f"gstp_{i}",
                label_visibility="collapsed",
            )

        # ── Safe numeric values for calculations ──────────────
        safe_qty   = int(qty)   if qty   is not None else 0
        safe_price = float(price) if price is not None else 0.0
        price_inc_gst         = safe_price * (1 + gst_p / 100)
        line_total_before_gst = safe_qty * safe_price
        line_total_inc_gst    = safe_qty * price_inc_gst

        # ── Total (display only) ──────────────────────────────
        with c7:
            total_display = f"₹{line_total_inc_gst:,.2f}" if (safe_qty > 0 or safe_price > 0) else ""
            st.text_input(
                f"Total {i}", value=total_display,
                disabled=True, label_visibility="collapsed",
            )

        # ── Delete button ─────────────────────────────────────
        with c8:
            st.markdown("<div style='padding-top:26px'>", unsafe_allow_html=True)
            if st.button("❌", key=f"remove_{i}"):
                items_to_remove.append(i)
            st.markdown("</div>", unsafe_allow_html=True)

        # ── Resolve final row identifiers ─────────────────────
        final_row_erp = "NEW" if is_new else row_erp
        final_row_vin = vin_input if is_new else vin_input

        # ── Sync session state with widget values ─────────────
        st.session_state.line_items[i] = {
            "erp_code":         final_row_erp,
            "vendor_item_no":   final_row_vin,
            "description":      desc_input,
            "qty":              qty,
            "price_before_gst": price,
            "gst_percent":      gst_p,
            "remarks":          live.get("remarks", ""),
            "is_new_item":      is_new,
        }

        # ── Validation: only on touched rows ──────────────────
        # (errors collected here shown only after submit attempt — see below)
        if row_touched:
            row_errors = []
            if not final_row_vin.strip():
                row_errors.append("Vendor Item No required")
            if safe_qty <= 0:
                row_errors.append("Qty must be > 0")
            if safe_price <= 0:
                row_errors.append("Price must be > 0")
            if row_errors:
                line_errors.append(f"Row {i+1}: " + " • ".join(row_errors))

            # Only accumulate into saving list when row is valid
            if not row_errors and final_row_vin.strip():
                line_data_for_saving.append({
                    "S.No":                  i + 1,
                    "ERP Code":              final_row_erp,
                    "Vendor Item No":        final_row_vin,
                    "Product Description":   desc_input,
                    "Qty":                   safe_qty,
                    "Price Before GST":      safe_price,
                    "GST %":                 gst_p,
                    "Price Inc. GST":        price_inc_gst,
                    "Remarks":               live.get("remarks", ""),
                    "line_total_before_gst": line_total_before_gst,
                    "line_total_inc_gst":    line_total_inc_gst,
                })

    # ── Show validation errors ONLY after user has clicked Submit ──
    if st.session_state.get("submit_validated") and line_errors:
        for err in line_errors:
            st.warning(f"⚠️ {err}")

    for idx in reversed(items_to_remove):
        st.session_state.line_items.pop(idx)
        st.rerun()

    # ── Totals ───────────────────────────────────────────────
    subtotal_before_gst = sum(l["line_total_before_gst"] for l in line_data_for_saving)
    total_inc_gst = sum(l["line_total_inc_gst"] for l in line_data_for_saving)
    gst_amt = total_inc_gst - subtotal_before_gst

    st.markdown("---")
    f_col1, f_col2 = st.columns([3, 1])
    with f_col2:
        freight_charge = st.number_input("Freight Charge", min_value=0.0, value=0.0, step=0.01)
    
    total_order_value = total_inc_gst + freight_charge

    t1, t2, t3, t4 = st.columns(4)
    t1.metric("Sub Total (Excl. GST)", f"₹{subtotal_before_gst:,.2f}")
    t2.metric("Total GST", f"₹{gst_amt:,.2f}")
    t3.metric("Freight Charge", f"₹{freight_charge:,.2f}")
    t4.metric("Total Order Value", f"₹{total_order_value:,.2f}")

    # ── Submit ───────────────────────────────────────────────
    st.markdown("---")
    
    update_reason = ""
    if edit_mode:
        update_reason = st.text_input("📝 Reason for Update / General Remarks", value="", placeholder="e.g., Price changed by vendor, added new items, etc.")
        st.markdown("<br>", unsafe_allow_html=True)

    sb1, sb2, sb3 = st.columns([1.2, 1.5, 4])
    with sb1:
        # Disable the button once quote has been submitted (prevents duplicates)
        already_submitted = st.session_state.get("quote_submitted", False)
        submit = st.button(
            "✅ Submitted" if already_submitted else "💾 Submit Quote",
            use_container_width=True,
            disabled=already_submitted,
        )
    with sb2:
        if already_submitted:
            st.success("Quote already submitted!")

    if submit and not st.session_state.get("quote_submitted", False):
        # Mark that the user has attempted submission — shows validation inline
        st.session_state.submit_validated = True

        errors = []
        if not salesperson.strip():
            errors.append("Salesperson Name is required.")
        if selected_vendor == "-- Select Vendor --":
            errors.append("Please select a Vendor.")
        if not line_data_for_saving:
            errors.append("Add at least one valid line item with Vendor Item No, Qty > 0 and Price > 0.")
        if line_errors:
            errors.extend(line_errors)

        if errors:
            for e in errors:
                st.error(e)
        else:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Fetch original Created Date and default Status in case it's not found
            orig_created = now
            orig_status = "Pending Approval"
            if edit_mode and not quotes_df.empty:
                existing_rows = quotes_df[quotes_df["Quote No"] == quote_no]
                if not existing_rows.empty:
                    orig_created = str(existing_rows.iloc[0].get("Created Date", now))
                    orig_status = str(existing_rows.iloc[0].get("Status", "Pending Approval"))

            # If in edit mode, the status must always be set to "Modified"
            target_status = "Modified" if edit_mode else "Pending Approval"
            modified_dt = now if edit_mode else ""

            final_rows = []
            for l in line_data_for_saving:
                # Combine specific line remarks with overall update reason
                specific_remark = l["Remarks"]
                if edit_mode and update_reason:
                    combined_remark = f"{specific_remark} | Update: {update_reason}" if specific_remark else update_reason
                else:
                    combined_remark = specific_remark
                
                final_rows.append({
                    "Quote No": quote_no,
                    "Quote Date": str(quote_date),
                    "Location Code": location_code,
                    "Purchase Quote raised by": salesperson.strip(),
                    "Vendor Name": selected_vendor,
                    "Vendor No": vendor_no,
                    "Vendor Address": vendor_addr,
                    "Vendor City": vendor_city,
                    "Vendor GST": vendor_gst,
                    "Vendor Email": vendor_email,
                    "Payment Method": payment_method,
                    "Cash": cash,
                    "Online Transfer": online_transfer,
                    "CDC": cdc,
                    "PDC": pdc,
                    "Credit Days": credit_days,
                    "Mode of Delivery": mode_of_delivery,
                    "Door Delivery": door_delivery,
                    "Pickup": pickup,
                    "Courier": courier,
                    "Expected Delivery Date": str(expected_delivery) if expected_delivery else "",
                    "Courier Details": courier_details,
                    "S.No": l["S.No"],
                    "ERP Code": l["ERP Code"],
                    "Vendor Item No": l["Vendor Item No"],
                    "Product Description": l["Product Description"],
                    "Qty": l["Qty"],
                    "Price Before GST": l["Price Before GST"],
                    "GST %": l["GST %"],
                    "Price Inc. GST": l["Price Inc. GST"],
                    "Total (Incl. GST)": l["line_total_inc_gst"],
                    "Remarks": combined_remark,
                    "Freight Charge": freight_charge,
                    "Status": target_status,
                    "Created Date": orig_created,
                    "Modified Date": modified_dt,
                })
            
            if final_rows:
                new_df = pd.DataFrame(final_rows)
                # Align columns to REQUIRED_COLS to prevent misaligned database inserts
                for col in REQUIRED_COLS:
                    if col not in new_df.columns:
                        new_df[col] = ""
                new_df = new_df[REQUIRED_COLS]
                
                if save_quotes(new_df, is_edit=edit_mode, quote_no=quote_no):
                    st.success(f"✅ Quote **{quote_no}** {'updated' if edit_mode else 'submitted'} successfully!")
                    st.session_state.submitted_quote_no = quote_no
                    # Lock submit button to prevent duplicates
                    st.session_state.quote_submitted  = True
                    st.session_state.submit_validated = False
                    st.cache_data.clear()
                    quotes_df = load_quotes()

        # ── Post-Submit / Edit Actions (Email & Download) ────────
    show_actions_for = quote_no if edit_mode else st.session_state.get("submitted_quote_no")
    if show_actions_for:
        st.markdown("---")
        st.subheader(f"Actions for Quote: {show_actions_for}")
        
        # We need the current rows to generate the email/PO
        q_rows_for_actions = quotes_df[quotes_df["Quote No"] == show_actions_for]
        
        c_dl, c_mail = st.columns([1.5, 2.5])
        
        with c_dl:
            # 1. Download Actions
            st.write("**📥 Export PO**")
            # Excel
            buf_xl = export_po_excel(show_actions_for, quotes_df)
            if buf_xl:
                st.download_button("Excel PO", buf_xl, file_name=f"PO_{show_actions_for}.xlsx", 
                                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 use_container_width=True)
            # PDF
            buf_pdf = export_po_pdf(show_actions_for, quotes_df)
            if buf_pdf:
                st.download_button("PDF PO", buf_pdf, file_name=f"PO_{show_actions_for}.pdf", 
                                 mime="application/pdf",
                                 use_container_width=True)
        
        with c_mail:
            # 2. Email Configuration
            st.write("**✉️ Email Submission**")
            total_qty_em = q_rows_for_actions["Qty"].fillna(0).sum()
            raw_fr_em = q_rows_for_actions["Freight Charge"].iloc[0] if "Freight Charge" in q_rows_for_actions.columns else 0
            fr_em = float(raw_fr_em) if pd.notnull(raw_fr_em) else 0.0
            line_totals = q_rows_for_actions["Qty"].fillna(0) * q_rows_for_actions["Price Inc. GST"].fillna(0)
            total_amt_em = line_totals.sum() + fr_em

            def_greeting = (
                "Dear Sir/Madam,\n\n"
                "Greetings from Supreme Computers India Pvt. Ltd.\n\n"
                "Please find below the Purchase Quote details for your reference:"
            )
            def_closing = (
                f"Total Qty: {int(total_qty_em)}\n"
                f"Total Amount (Incl. GST): {total_amt_em:,.2f}\n\n"
                "Kindly review and confirm. "
                "Please feel free to contact us for any clarification."
            )

            email_greeting = st.text_area("Email Greeting & Intro", value=def_greeting, height=120)
            email_closing = st.text_area("Email Closing & Totals", value=def_closing, height=120)

            subject_em = f"Purchase Quote Submission - {show_actions_for}"
            cc_emails = "purchase@supremeindia.com, mis3@supremeindia.com"

            # ── Build Outlook-Classic-safe HTML using list, not f-string loop ──
            # Keeps every tag on a full line, avoids QP line-break corruption.
            def _safe(v):
                """Escape special HTML chars; return plain string (no f-string tags)."""
                import html as _html
                return _html.escape(str(v) if pd.notnull(v) else "")

            TD  = 'style="border:1px solid #000000;padding:5px 8px;font-family:Calibri,Arial,sans-serif;font-size:13px;"'
            TDC = 'style="border:1px solid #000000;padding:5px 8px;text-align:center;font-family:Calibri,Arial,sans-serif;font-size:13px;"'
            TDR = 'style="border:1px solid #000000;padding:5px 8px;text-align:right;font-family:Calibri,Arial,sans-serif;font-size:13px;"'
            HDR = 'style="border:1px solid #000000;padding:5px 8px;background-color:#BDD7EE;font-weight:bold;text-align:center;font-family:Calibri,Arial,sans-serif;font-size:13px;"'
            TOT = 'style="border:1px solid #000000;padding:5px 8px;text-align:right;background-color:#f2f2f2;font-weight:bold;font-family:Calibri,Arial,sans-serif;font-size:13px;"'
            TOTL= 'style="border:1px solid #000000;padding:5px 8px;text-align:center;background-color:#E2EFDA;font-weight:bold;font-family:Calibri,Arial,sans-serif;font-size:13px;"'

            # Header row
            html_parts = [
                '<table border="1" cellspacing="0" cellpadding="0" '
                'style="border-collapse:collapse;width:100%;border:1px solid #000000;">',
                '<tr>',
                '<td ' + HDR + '>S.No</td>',
                '<td ' + HDR + '>Vendor Item No</td>',
                '<td ' + HDR + '>Product Description</td>',
                '<td ' + HDR + '>Qty</td>',
                '<td ' + HDR + '>Price (Excl.)</td>',
                '<td ' + HDR + '>GST %</td>',
                '<td ' + HDR + '>Total (Incl. GST)</td>',
                '</tr>',
            ]

            # Data rows — build each cell individually, never embed tags in f-string values
            for _, r in q_rows_for_actions.iterrows():
                qty    = float(r['Qty'])           if pd.notnull(r['Qty'])           else 0.0
                p_b    = float(r['Price Before GST']) if pd.notnull(r['Price Before GST']) else 0.0
                g_p    = float(r['GST %'])          if pd.notnull(r['GST %'])          else 0.0
                p_i    = float(r['Price Inc. GST']) if pd.notnull(r['Price Inc. GST']) else 0.0
                row_total = qty * p_i

                sno_val  = _safe(r['S.No'])
                vin_val  = _safe(r['Vendor Item No'])
                desc_val = _safe(r['Product Description'])
                qty_val  = str(int(qty))
                pb_val   = "{:,.2f}".format(p_b)
                gp_val   = str(g_p)
                rt_val   = "{:,.2f}".format(row_total)

                html_parts += [
                    '<tr>',
                    '<td ' + TDC + '>' + sno_val  + '</td>',
                    '<td ' + TD  + '>' + vin_val  + '</td>',
                    '<td ' + TD  + '>' + desc_val + '</td>',
                    '<td ' + TDC + '>' + qty_val  + '</td>',
                    '<td ' + TDR + '>' + pb_val   + '</td>',
                    '<td ' + TDC + '>' + gp_val   + '</td>',
                    '<td ' + TDR + '>' + rt_val   + '</td>',
                    '</tr>',
                ]

            # Totals row
            tqty_val = str(int(total_qty_em))
            tamt_val = "{:,.2f}".format(total_amt_em)
            html_parts += [
                '<tr>',
                '<td colspan="3" ' + TD  + '></td>',
                '<td ' + TOT + '>' + tqty_val + '</td>',
                '<td ' + TD  + '></td>',
                '<td ' + TOTL+ '>Total</td>',
                '<td ' + TOT + '>' + tamt_val + '</td>',
                '</tr>',
                '</table>',
            ]

            # Join with newlines — keeps lines short, avoids QP wrapping mid-tag
            html_table = "\n".join(html_parts)

            # ── Plain-text fallback (tab-separated) ──
            txt_lines = ["S.No\tVendor Item No\tProduct Description\tQty\tPrice (Excl.)\tGST %\tTotal (Incl.)"]
            for _, r in q_rows_for_actions.iterrows():
                qty = float(r['Qty'])           if pd.notnull(r['Qty'])           else 0.0
                p_b = float(r['Price Before GST']) if pd.notnull(r['Price Before GST']) else 0.0
                g_p = float(r['GST %'])          if pd.notnull(r['GST %'])          else 0.0
                p_i = float(r['Price Inc. GST']) if pd.notnull(r['Price Inc. GST']) else 0.0
                row_total = qty * p_i
                txt_lines.append(
                    "\t".join([
                        str(r['S.No']),
                        str(r['Vendor Item No']),
                        str(r['Product Description']),
                        str(int(qty)),
                        "{:,.2f}".format(p_b),
                        str(g_p),
                        "{:,.2f}".format(row_total),
                    ])
                )
            text_table = "\n".join(txt_lines)

            # ── Build full HTML body ──
            greeting_html = email_greeting.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
            closing_html  = email_closing.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")

            html_body = "\n".join([
                "<!DOCTYPE html>",
                "<html>",
                "<head>",
                '<meta http-equiv="Content-Type" content="text/html; charset=UTF-8">',
                "</head>",
                "<body>",
                '<p style="font-family:Calibri,Arial,sans-serif;font-size:13px;">',
                greeting_html,
                "</p>",
                html_table,
                '<p style="font-family:Calibri,Arial,sans-serif;font-size:13px;">',
                closing_html,
                "</p>",
                "</body>",
                "</html>",
            ])

            # ── Assemble .eml using MIMEMultipart — avoids QP on HTML part ──
            # Using email.mime directly gives full control over Content-Transfer-Encoding.
            import email.mime.multipart as _mp
            import email.mime.text      as _mt
            import email.mime.base      as _mb
            import email.mime.application as _ma
            from email import encoders as _enc

            outer = _mp.MIMEMultipart("mixed")
            outer["Subject"] = subject_em
            outer["Cc"]      = cc_emails
            outer["MIME-Version"] = "1.0"

            # multipart/alternative holds text + html
            alt = _mp.MIMEMultipart("alternative")

            # Plain-text part — utf-8, quoted-printable is fine here (plain text)
            plain_body = email_greeting + "\n\n" + text_table + "\n\n" + email_closing
            alt.attach(_mt.MIMEText(plain_body, "plain", "utf-8"))

            # HTML part — force base64 so QP encoding NEVER corrupts the tags
            html_part = _mt.MIMEText(html_body, "html", "utf-8")
            # Override the default QP transfer encoding with base64
            html_encoded = html_body.encode("utf-8")
            import base64 as _b64
            html_part_b64 = _mb.MIMEBase("text", "html")
            html_part_b64["Content-Type"] = 'text/html; charset="utf-8"'
            html_part_b64["Content-Transfer-Encoding"] = "base64"
            html_part_b64.set_payload(_b64.encodebytes(html_encoded).decode("ascii"))
            alt.attach(html_part_b64)

            outer.attach(alt)

            # PDF attachment
            if buf_pdf:
                pdf_bytes = buf_pdf.getvalue()
                pdf_part = _mb.MIMEBase("application", "pdf")
                pdf_part["Content-Disposition"] = f'attachment; filename="PO_{show_actions_for}.pdf"'
                pdf_part["Content-Transfer-Encoding"] = "base64"
                pdf_part.set_payload(_b64.encodebytes(pdf_bytes).decode("ascii"))
                outer.attach(pdf_part)

            eml_bytes = outer.as_bytes()

            st.download_button(
                "🚀 Open Outlook Draft (with PO PDF)",
                eml_bytes,
                file_name=f"Draft_{show_actions_for}.eml",
                mime="message/rfc822",
                use_container_width=True,
            )
            st.success(
                "✅ **Draft Ready!** Click above to download the Outlook file. "
                "Your **PO PDF will be automatically attached** and the table "
                "will render correctly in Gmail, Outlook Web, and Outlook Classic."
            )

        # ── Show Table for Preview ──
        st.markdown(html_table, unsafe_allow_html=True)

# ── VIEW QUOTES ──────────────────────────────────────────────
def page_view_quotes():
    st.markdown('<div class="section-header">🔍 View Quotes</div>', unsafe_allow_html=True)
    df = load_quotes()
    if df.empty:
        st.info("No quotes available yet.")
        return

    # Filters
    f1, f2, f3 = st.columns(3)
    with f1:
        sp_list = ["All"] + df["Purchase Quote raised by"].dropna().unique().tolist()
        sp_filter = st.selectbox("Filter by Purchaser", sp_list)
    with f2:
        st_list = ["All"] + df["Status"].dropna().unique().tolist()
        st_filter = st.selectbox("Filter by Status", st_list)
    with f3:
        date_filter = st.date_input("Filter by Date (Created)", value=None)

    filtered = df.copy()
    if sp_filter != "All":
        filtered = filtered[filtered["Purchase Quote raised by"] == sp_filter]
    if st_filter != "All":
        filtered = filtered[filtered["Status"] == st_filter]
    if date_filter:
        filtered["_date"] = pd.to_datetime(filtered["Created Date"], errors="coerce").dt.date
        filtered = filtered[filtered["_date"] == date_filter]
        filtered = filtered.drop(columns=["_date"], errors="ignore")

    st.markdown(f"**Showing {len(filtered)} records**")
    disp_f = filtered.copy()
    for c in disp_f.select_dtypes(include='object').columns:
        disp_f[c] = disp_f[c].astype(str)
    st.dataframe(disp_f, width='stretch')

    # Download
    buf = io.BytesIO()
    filtered.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    st.download_button("⬇️ Download as Excel", buf, file_name="quotes_export.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=False)

# ── Routing ──────────────────────────────────────────────────
page = st.session_state.page
if page == "Dashboard":
    page_dashboard()
elif page == "Create / Edit Quote":
    page_create_quote()
elif page == "View Quotes":
    page_view_quotes()
elif page == "Manage Quotes":
    st.warning("The Manage menu has been moved inside the 'Create / Edit Quote' page. Please use the dropdown at the top of that page to select a quote to manage.")
